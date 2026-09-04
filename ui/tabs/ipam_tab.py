import io
import re
import ipaddress
import streamlit as st
import pandas as pd
import openpyxl
from core.ipam_engine import (
    VLAN_PRESETS,
    BRANCH_VLAN_PRESET,
    DATACENTER_VLAN_PRESET,
    compute_chained_rows,
    slugify,
    evaluate_subnet_row,
    calculate_remaining_subnets,
    get_subnet_availability_analysis,
    calculate_ip_range_str,
    format_branch_display,
    lookup_role_description,
    generate_netbox_site_csv,
    generate_netbox_vlan_group_csv,
    generate_netbox_vlans_csv,
    generate_netbox_prefixes_csv
)
from core.db_manager import (get_all_site_names,
    save_ipam_records_batch,
    save_sites_batch,
    clear_ipam_records,
    clear_sites_records,
    clear_vlans_records,
    clear_prefixes_records,
    get_total_ipam_count,
    get_total_sites_count,
    get_total_vlans_count,
    get_total_prefixes_count,
    get_max_scope_id,
    lookup_scope_id,
    lookup_site_supernet_from_db,
    get_existing_prefix_strings,
    get_sync_metadata,
    get_site_summary,
    get_ipam_records_by_site,
    get_full_site_inventory_summary
)

from utils.formatters import to_title_case_preserve_acronyms
from ui.components import render_ai_chat, render_backup_uploader

def handle_site_change():
    """Triggered on site name input change: automatically looks up and fills Scope ID & Supernet."""
    entered_site = st.session_state.get("ipam_site_in", "").strip()
    if entered_site:
        matched_scope = lookup_scope_id(entered_site)
        matched_super = lookup_site_supernet_from_db(entered_site)
        if matched_scope is not None:
            st.session_state["ipam_scope_in"] = str(matched_scope)
        if matched_super is not None:
            st.session_state["ipam_super_in"] = str(matched_super)

def handle_ipam_file_upload():
    uploader_key_val = st.session_state.get("uploader_key", 0)
    uploaded_files = st.session_state.get(f"ipam_multi_uploader_{uploader_key_val}")
    if not uploaded_files:
        return

    if not isinstance(uploaded_files, list):
        uploaded_files = [uploaded_files]

    total_scopes = 0
    total_prefixes = 0
    errors = []

    for file_obj in uploaded_files:
        filename = file_obj.name.lower()
        content = file_obj.getvalue()
        
        if filename.endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                scope_records = []
                ipam_records = []

                if "Scope" in wb.sheetnames:
                    ws_scope = wb["Scope"]
                    for r in range(2, ws_scope.max_row + 1):
                        s_id = ws_scope.cell(row=r, column=1).value
                        s_name = ws_scope.cell(row=r, column=5).value
                        s_slug = ws_scope.cell(row=r, column=6).value
                        if s_name:
                            scope_records.append({"id": s_id, "name": s_name, "slug": s_slug})
                    if scope_records:
                        cnt = save_sites_batch(scope_records, clear_first=False, source="Manual CSV Upload")
                        total_scopes += cnt

                if "Prefixes" in wb.sheetnames:
                    ws_pfx = wb["Prefixes"]
                    for r in range(2, ws_pfx.max_row + 1):
                        pfx_str = ws_pfx.cell(row=r, column=6).value
                        scope_id_val = ws_pfx.cell(row=r, column=9).value
                        vlan_val = ws_pfx.cell(row=r, column=12).value
                        role_val = ws_pfx.cell(row=r, column=14).value
                        desc_val = ws_pfx.cell(row=r, column=17).value
                        if pfx_str and str(pfx_str).strip():
                            ipam_records.append({
                                "prefix_or_subnet": str(pfx_str).strip(),
                                "scope_id": scope_id_val,
                                "vlan_name": str(vlan_val or ""),
                                "role": str(role_val or ""),
                                "description": str(desc_val or "")
                            })
                    if ipam_records:
                        cnt = save_ipam_records_batch(ipam_records, clear_first=False, source="Manual CSV Upload")
                        total_prefixes += cnt
            except Exception as e:
                errors.append(f"• **{file_obj.name}**: {str(e)}")
        else:
            try:
                # Use engine='python' and on_bad_lines='skip' for more robust parsing
                # Also try to detect delimiter
                import io
                content_str = content.decode('utf-8', errors='replace')
                first_line = content_str.splitlines()[0] if content_str else ""
                delim = ','
                if ';' in first_line and first_line.count(';') > first_line.count(','):
                    delim = ';'
                
                df = pd.read_csv(io.StringIO(content_str), sep=delim)
                cols = {str(c).lower().strip(): c for c in df.columns}
                st.toast(f"DEBUG: CSV Columns={list(cols.keys())}", icon="🔍")

                # Robust site format detection: distinguish from prefixes/VLANs

                is_site_file = ("name" in cols or "site" in cols or "location" in cols) and not ("prefix" in cols or "prefixes" in cols or "vid" in cols or "vlan" in cols)
                if is_site_file:
                    name_col = cols.get("name", cols.get("site", cols.get("location")))
                    id_col = cols.get("id")
                    slug_col = cols.get("slug")
                    scope_records = []
                    for idx, row in df.iterrows():
                        s_name = str(row.get(name_col, "")).strip()
                        s_id = row.get(id_col) if id_col and not pd.isna(row.get(id_col)) else (idx + 1)
                        s_slug = str(row.get(slug_col, "")).strip()
                        
                        # Generate slug if missing
                        if not s_slug and s_name:
                            s_slug = slugify(s_name)
                            
                        if s_name and s_name.lower() != "nan":
                            scope_records.append({"id": int(s_id) if str(s_id).isdigit() else s_id, "name": s_name, "slug": s_slug})
                    if scope_records:
                        cnt = save_sites_batch(scope_records, clear_first=False, source="Manual CSV Upload")
                        total_scopes += cnt
                elif "prefixes" in cols or "prefix" in cols or "subnet" in cols or "vid" in cols or "vlan" in cols:
                    st.toast(f"DEBUG: CSV Columns={list(cols.keys())}", icon="🔍")
                    st.session_state["ipam_multi_uploader"] = None
                    pfx_col = cols.get("prefixes", cols.get("prefix", cols.get("subnet")))
                    vid_col = cols.get("vid", cols.get("vlan_id", cols.get("vlan", "")))
                    vname_col = cols.get("vlan_name", cols.get("vlan name", cols.get("name", "")))
                    site_col = cols.get("site", cols.get("site name", cols.get("location", cols.get("scope", ""))))
                    role_col = cols.get("role", cols.get("role name", ""))
                    desc_col = cols.get("description", cols.get("comments", cols.get("desc", "")))

                    is_vlan_file = "vlan" in filename or "vlans" in filename or ("vid" in cols and not ("prefixes" in filename or "prefix" in filename))
                    rec_type = "vlan" if is_vlan_file else "prefix"

                    ipam_records = []
                    # Robust CSV import: handle Pandas NaN correctly
                    for _, row in df.iterrows():
                        raw_prefixes = str(row.get(pfx_col, "")).strip() if pfx_col and not pd.isna(row.get(pfx_col)) else ""
                        vid = str(row.get(vid_col, "")).strip() if vid_col and not pd.isna(row.get(vid_col)) else ""
                        vname = str(row.get(vname_col, "")).strip() if vname_col and not pd.isna(row.get(vname_col)) else ""
                        role_str = str(row.get(role_col, "")).strip() if role_col and not pd.isna(row.get(role_col)) else ""
                        site_str = str(row.get(site_col, "")).strip() if site_col and not pd.isna(row.get(site_col)) else ""
                        desc_str = str(row.get(desc_col, "")).strip() if desc_col and not pd.isna(row.get(desc_col)) else ""

                        if not raw_prefixes or raw_prefixes.lower() == "nan":
                            if rec_type == "vlan":
                                if vid or vname:
                                    ipam_records.append({
                                        "prefix_or_subnet": "",
                                        "vlan_id": vid if vid.isdigit() else None,
                                        "vlan_name": vname if vname.lower() != "nan" else "",
                                        "role": role_str if role_str.lower() != "nan" else "",
                                        "site": site_str if site_str.lower() != "nan" else "",
                                        "description": desc_str if desc_str.lower() != "nan" else "",
                                        "record_type": "vlan"
                                    })
                            continue

                        found_cidrs = re.findall(r'\b(?:\d{1,3}\.){3}\d{1,3}/\d{1,2}\b', raw_prefixes)

                        if found_cidrs:
                            for cidr in found_cidrs:
                                ipam_records.append({
                                    "prefix_or_subnet": cidr,
                                    "vlan_id": vid if vid.isdigit() else None,
                                    "vlan_name": vname if vname.lower() != "nan" else "",
                                    "role": role_str if role_str.lower() != "nan" else "",
                                    "site": site_str if site_str.lower() != "nan" else "",
                                    "description": desc_str if desc_str.lower() != "nan" else "",
                                    "record_type": rec_type
                                })
                        elif "/" in raw_prefixes:
                            ipam_records.append({
                                "prefix_or_subnet": raw_prefixes,
                                "vlan_id": vid if vid.isdigit() else None,
                                "vlan_name": vname if vname.lower() != "nan" else "",
                                "role": role_str if role_str.lower() != "nan" else "",
                                "site": site_str if site_str.lower() != "nan" else "",
                                "description": desc_str if desc_str.lower() != "nan" else "",
                                "record_type": rec_type
                            })

                    if ipam_records:
                        cnt = save_ipam_records_batch(ipam_records, clear_first=False, source="Manual CSV Upload")
                        total_prefixes += cnt
                    
                    st.session_state["ipam_multi_uploader"] = None
                else:
                    raise ValueError(f"Unrecognized CSV format. Found columns: {list(cols.keys())}. Expected `netbox_sites.csv` columns (slug, name, id) or `netbox_prefixes.csv`/`netbox_VLANs.csv` columns (prefix, vlan, vid).")
            except Exception as e:
                errors.append(f"• **{file_obj.name}**: {str(e)}")

    if errors:
        for err in errors:
            st.error(err)

        if total_scopes > 0 or total_prefixes > 0:
            st.toast(f"✅ Ingested: {total_scopes} Sites, {total_prefixes} Prefixes!", icon="🚀")
            st.session_state["ipam_multi_uploader"] = []
            st.rerun()


def handle_ipam_db_reset():
    clear_ipam_records()
    if "ipam_persisted_rows" in st.session_state:
        del st.session_state["ipam_persisted_rows"]
    if "ipam_scope_in" in st.session_state:
        del st.session_state["ipam_scope_in"]
    if "ipam_site_in" in st.session_state:
        del st.session_state["ipam_site_in"]
    if "ipam_super_in" in st.session_state:
        del st.session_state["ipam_super_in"]
    st.toast("🗑️ Database Cleared. Restored default templates.", icon="🧹")

def on_preset_change():
    selected = st.session_state.get("ipam_preset_selector")
    site_name = st.session_state.get("ipam_site_in", "").strip()
    
    # Reset site found indicator
    st.session_state["ipam_site_found_in_db"] = False

    if selected == "🗄️ Load From DB (Existing Site)" and site_name:
        records = get_ipam_records_by_site(site_name)
        
        if not records:
            st.session_state["ipam_persisted_rows"] = []
            st.session_state["ipam_site_found_in_db"] = False
            return
        
        # Mark that site was found in DB
        st.session_state["ipam_site_found_in_db"] = True
        
        # Determine if it's a Branch or DC to apply sorting
        branch_vids = {p['vid'] for p in BRANCH_VLAN_PRESET}
        dc_vids = {p['vid'] for p in DATACENTER_VLAN_PRESET}
        
        match_branch = sum(1 for r in records if r.get('vlan_id') in branch_vids)
        match_dc = sum(1 for r in records if r.get('vlan_id') in dc_vids)
        
        target_preset = BRANCH_VLAN_PRESET if match_branch >= match_dc else DATACENTER_VLAN_PRESET
        order_map = {p['vid']: i for i, p in enumerate(target_preset)}
        
        # Sort records by order in preset, then by vlan_id (handle None values)
        def sort_key(r):
            vlan_id = r.get('vlan_id')
            # If vlan_id is None, treat it as 9999 for sorting (put at end)
            if vlan_id is None:
                return (999, 9999)
            return (order_map.get(vlan_id, 999), vlan_id)
        
        records.sort(key=sort_key)
        
        # Use a dictionary to de-duplicate based on vlan_id
        # Prefer records with non-empty role, vlan_name, and description
        unique_records = {}
        for r in records:
            vid = r.get("vlan_id")
            if vid is None:
                continue
                
            # If this VLAN ID hasn't been seen yet, add it
            if vid not in unique_records:
                unique_records[vid] = r
            else:
                # Already have a record for this VLAN - keep the better one
                existing = unique_records[vid]
                
                # Calculate "quality score" - prefer records with more filled fields
                def quality_score(rec):
                    score = 0
                    if rec.get("role", "").strip():
                        score += 3  # Role is most important
                    if rec.get("description", "").strip():
                        score += 2
                    if rec.get("vlan_name", "").strip():
                        score += 1
                    if rec.get("prefix_or_subnet", "").strip():
                        score += 1
                    return score
                
                if quality_score(r) > quality_score(existing):
                    unique_records[vid] = r
                
        new_rows = []
        for vid, r in unique_records.items():
            new_rows.append({
                "VLAN ID": vid,
                "Role": r.get("role", ""),
                "VLAN Name": r.get("vlan_name", ""),
                "VLAN Description": r.get("description", ""),
                "Subnet (CIDR)": r.get("prefix_or_subnet", "")
            })
        st.session_state["ipam_persisted_rows"] = new_rows
        return

    template_list = VLAN_PRESETS.get(selected, [])
    
    if "ipam_data_editor_live" in st.session_state:
        del st.session_state["ipam_data_editor_live"]

    if not template_list:
        st.session_state["ipam_persisted_rows"] = []
    else:
        new_rows = []
        for t in template_list:
            role_name = t["role"]
            new_rows.append({
                "VLAN ID": t["vid"],
                "Role": role_name,
                "VLAN Name": t.get("vlan_name", role_name),
                "VLAN Description": t.get("desc", lookup_role_description(role_name)),
                "Subnet (CIDR)": ""
            })
        st.session_state["ipam_persisted_rows"] = new_rows

def build_ipam_system_prompt(prompt, site_name, supernet_in, existing_prefixes, max_scope_id):
    """Build the grounded IPAM system prompt for the AI Assistant."""
    from core.ai_helper import build_comprehensive_ipam_context

    comprehensive_context = build_comprehensive_ipam_context(prompt, site_filter=site_name)

    # Combine DB prefixes with the prefixes currently in the allocation editor
    ui_prefixes = []
    for r in st.session_state.get("ipam_persisted_rows", []):
        sub = str(r.get("Subnet (CIDR)", "") or "").strip()
        if sub and "/" in sub:
            ui_prefixes.append(sub)

    combined_prefixes = list(dict.fromkeys(existing_prefixes + ui_prefixes))

    # Extract requested mask (e.g., /24) from user prompt if specified
    mask_match = re.search(r"/(\d{1,2})", prompt)
    req_mask = int(mask_match.group(1)) if mask_match else 24

    # Determine target supernet for pre-calculation
    target_networks = []
    cidr_matches = re.findall(r"\b(?:\d{1,3}\.){3}\d{1,3}(?:/\d{1,2})?\b", prompt)
    if supernet_in:
        cidr_matches.append(supernet_in)

    for match in cidr_matches:
        try:
            target_networks.append(ipaddress.ip_network(match, strict=False))
        except Exception:
            pass

    calc_supernet = supernet_in
    if not calc_supernet and target_networks:
        calc_supernet = str(target_networks[0])

    calc_analysis = ""
    if calc_supernet and "/" in calc_supernet:
        calc_analysis = get_subnet_availability_analysis(
            calc_supernet,
            combined_prefixes,
            requested_prefix_len=req_mask
        )

    next_scope_id_val = (max_scope_id + 1) if max_scope_id is not None else 1

    return f"""You are an expert network architect specializing in IP address management, NetBox provisioning, and subnet planning.
You have DIRECT ACCESS to the complete IPAM database. Analyze the user's request and respond accurately using the ACTUAL DATABASE DATA provided below.

=== COMPLETE DATABASE CONTEXT ===
{comprehensive_context}

=== CURRENT UI STATE ===
- Current Site Input: {site_name or 'Not specified'}
- Current Supernet Input: {supernet_in or 'Not specified'}
- Next Available Scope ID: {next_scope_id_val}
- Max Scope ID in DB: {max_scope_id or 'None'}

{calc_analysis}

=== IMPORTANT GUIDELINES ===
1. **Answer ALL questions using the ACTUAL DATA above** - You have complete database access
2. When asked about VLANs, prefixes, or subnets for a site:
   - Reference the "VLANs and Prefixes for Site" section above
   - List the REAL VLAN IDs, names, subnets, roles, and descriptions
   - Be specific with actual data, not generic examples
3. When asked about devices or inventory:
   - Reference the "Inventory for Site" section if available
   - Provide actual device names, counts, and details
4. For subnet suggestions:
   - Follow the PRE-CALCULATED SUBNET AVAILABILITY ANALYSIS
   - Never suggest subnets marked as OCCUPIED or OVERLAPS
   - Explain why suggestions are available
5. For Scope ID questions:
   - Use the "Next Available Scope ID" value
6. Format responses clearly:
   - Use bullet points for lists
   - Include VLAN IDs, names, and subnets
   - Show actual counts and statistics
7. If no data exists for a query, clearly state "No records found in database"
8. Be concise but comprehensive - show all relevant data
9. When counting items (e.g., "how many VLANs"), provide the exact number from the data above
10. If a "NETBOX MASTER BACKUP" section is present it is the full NetBox export:
   - Treat it as authoritative for any object type (sites, racks, devices, interfaces, IPs, VMs, clusters, tenants, circuits, VRFs)
   - Use its "total in backup" figures when stating counts
   - Quote the exact records listed rather than generalizing"""

def render_ipam_tab(active_model: str):
    st.subheader("🌐 IPAM & Site Subnet Provisioning Engine")
    st.caption("Plan site supernets, allocate non-overlapping VLAN subnets, and export ready-to-import NetBox CSV blocks.")

    # 1. Ingestion Toolbar
    total_ipam_recs = get_total_ipam_count()
    total_sites_recs = get_total_sites_count()
    total_vlans_recs = get_total_vlans_count()
    total_prefixes_recs = get_total_prefixes_count()
    total_db_count = total_ipam_recs + total_sites_recs
    meta = get_sync_metadata("ipam")

    status_tag = f"🟢 ({total_sites_recs} sites, {total_vlans_recs} VLANs, {total_prefixes_recs} prefixes in DB)" if total_db_count > 0 else "⚪ (Default Examples)"
    tick_sites = " ✅" if total_sites_recs > 0 else ""
    tick_vlans = " ✅" if total_vlans_recs > 0 else ""
    tick_prefixes = " ✅" if total_prefixes_recs > 0 else ""

    with st.expander(f"📥 Ingest NetBox Site Data (Backup / CSV) {status_tag}", expanded=False):
        if total_db_count > 0:
            # Get the most recent source from any of the files
            meta_sites = get_sync_metadata("netbox_sites")
            meta_vlans = get_sync_metadata("netbox_VLANs")
            meta_prefixes = get_sync_metadata("netbox_prefixes")
            
            # Use the most recent source that's not "None"
            sources = [m['source'] for m in [meta_sites, meta_vlans, meta_prefixes] if m['source'] != "None"]
            display_source = sources[0] if sources else "Manual CSV Upload"
            
            st.markdown(f"**DB Status:** `Source: {display_source}`")

        render_backup_uploader("ipam")
        st.markdown("---")

        c_ref_row, c_ref_cap = st.columns([1, 3])
        with c_ref_row:
            if st.button("🔄 Refresh", key="ref_ipam_btn", width="stretch"):
                st.rerun()
        with c_ref_cap:
            st.caption("Reload the local database view.")

        st.markdown("**Option B: Manual CSV Export & Upload:**")

        meta_sites = get_sync_metadata("netbox_sites")
        meta_vlans = get_sync_metadata("netbox_VLANs")
        meta_prefixes = get_sync_metadata("netbox_prefixes")
        
        sites_timestamp = f" `{meta_sites['updated_at']}`" if meta_sites['updated_at'] != "Never" else ""
        vlans_timestamp = f" `{meta_vlans['updated_at']}`" if meta_vlans['updated_at'] != "Never" else ""
        prefixes_timestamp = f" `{meta_prefixes['updated_at']}`" if meta_prefixes['updated_at'] != "Never" else ""

        col_l1, col_r1 = st.columns([12, 1])
        with col_l1:
            st.markdown(f"* **Scope IDs & Site Names:** Go to `Organization` ➔ `Sites` ➔ `Export` ➔ `All Data` (`netbox_sites.csv`){tick_sites}{sites_timestamp}")
        with col_r1:
            if total_sites_recs > 0:
                if st.button("🗑️", key="btn_clr_sites_inline", help="Clear netbox_sites.csv data"):
                    clear_sites_records()
                    st.toast("🗑️ Cleared netbox_sites.csv data.", icon="🧹")
                    st.rerun()

        col_l2, col_r2 = st.columns([12, 1])
        with col_l2:
            st.markdown(f"* **VLANs:** Go to `IPAM` ➔ `VLANs` ➔ `Export` ➔ `All Data` (`netbox_VLANs.csv`){tick_vlans}{vlans_timestamp}")
        with col_r2:
            if total_vlans_recs > 0:
                if st.button("🗑️", key="btn_clr_vlans_inline", help="Clear netbox_VLANs.csv data"):
                    clear_vlans_records()
                    st.toast("🗑️ Cleared netbox_VLANs.csv data.", icon="🧹")
                    st.session_state["ipam_multi_uploader"] = None
                    st.rerun()

        col_l3, col_r3 = st.columns([12, 1])
        with col_l3:
            st.markdown(f"* **IP Prefixes:** Go to `IPAM` ➔ `Prefixes` ➔ `Export` ➔ `All Data` (`netbox_prefixes.csv`){tick_prefixes}{prefixes_timestamp}")
        with col_r3:
            if total_prefixes_recs > 0:
                if st.button("🗑️", key="btn_clr_prefixes_inline", help="Clear netbox_prefixes.csv data"):
                    clear_prefixes_records()
                    st.toast("🗑️ Cleared netbox_prefixes.csv data.", icon="🧹")
                    st.session_state["ipam_multi_uploader"] = None
                    st.rerun()

        c_up, c_rst = st.columns([3, 1])
        with c_up:
            st.file_uploader(
                "Upload NetBox CSVs (netbox_sites.csv, netbox_VLANs.csv, netbox_prefixes.csv) or Excel", 
                type=["xlsx", "csv"], 
                accept_multiple_files=True,
                key=f"ipam_multi_uploader_{st.session_state.get('uploader_key', 0)}",
                on_change=handle_ipam_file_upload,
                label_visibility="collapsed"
            )

        with c_rst:
            if total_db_count > 0:
                st.button("🗑️ Clear All DB", on_click=handle_ipam_db_reset, width="stretch", key="rst_ipam_csv_btn")
            else:
                st.caption("No custom data loaded.")

    # 2. Site Inputs and Dynamic Lookups
    if "ipam_site_in" not in st.session_state:
        st.session_state["ipam_site_in"] = ""
    if "ipam_scope_in" not in st.session_state:
        st.session_state["ipam_scope_in"] = ""
    if "ipam_super_in" not in st.session_state:
        st.session_state["ipam_super_in"] = ""

    # Get variables needed for AI Assistant
    max_scope_id = get_max_scope_id()
    scope_eg = f"e.g. {max_scope_id + 1}" if max_scope_id is not None else "e.g. 42"
    site_name = st.session_state.get("ipam_site_in", "").strip()
    supernet_in = st.session_state.get("ipam_super_in", "").strip()
    existing_prefixes = get_existing_prefix_strings()

    # 2.5. AI IPAM & Subnet Assistant
    render_ai_chat(
        history_key="ipam_chat_history",
        caption="Ask for subnet suggestions using natural language (e.g., 'Show me VLANs in Weybridge' or 'Suggest the next available /24 in 10.113.0.0/16')",
        placeholder="Ask for subnet suggestions...",
        active_model=active_model,
        build_system_prompt=lambda p: build_ipam_system_prompt(
            p, site_name, supernet_in, existing_prefixes, max_scope_id
        ),
    )

    # Site input fields
    # Site input fields
    top1, top2, top3 = st.columns([2, 1, 2.2])
    with top1:
        st.text_input(
            "Branch / Site Name",
            key="ipam_site_in",
            placeholder="e.g. Bristol, AGE, Adelaide, UK, Site-01",
            on_change=handle_site_change
        )

    auto_scope_id = lookup_scope_id(site_name) if site_name else None
    auto_supernet = lookup_site_supernet_from_db(site_name) if site_name else None

    if site_name and auto_supernet and not st.session_state.get("ipam_super_in"):
        st.session_state["ipam_super_in"] = str(auto_supernet)
        st.rerun()

    with top2:
        st.text_input(
            "Scope ID (NetBox Site ID)", 
            key="ipam_scope_in",
            placeholder=scope_eg,
            help="Auto-discovered from uploaded data/agent sync, or editable manually."
        )
        scope_id = st.session_state["ipam_scope_in"].strip()
        if auto_scope_id:
            st.caption(f"🟢 Matched Scope ID: **`{auto_scope_id}`**")
        else:
            st.caption("⚪ Manual Scope ID mode")

    with top3:
        st.text_input(
            "Site Supernet (CIDR)", 
            key="ipam_super_in",
            placeholder="e.g. 10.x.x.0/24",
            help="Top-level container subnet for this branch site."
        )
        cap_placeholder = st.empty()

    display_site_name = format_branch_display(site_name)

    # 3. Preset Selection & Allocation Editor
    st.markdown("---")
    c_title, c_preset, c_indicator = st.columns([2.2, 1.5, 0.5])
    with c_title:
        st.markdown("##### 📊 Subnet Allocation & Live Status (✏️ Click any cell to edit)")
    with c_preset:
        st.selectbox(
            "Load Standard Preset",
            options=list(VLAN_PRESETS.keys()),
            index=0,
            key="ipam_preset_selector",
            on_change=on_preset_change,
            help="Quickly load pre-defined standard VLAN structures or start blank."
        )
    with c_indicator:
        # Show checkbox indicator when "Load From DB" is selected and site was found
        selected_preset = st.session_state.get("ipam_preset_selector", "")
        site_found = st.session_state.get("ipam_site_found_in_db", False)
        
        if selected_preset == "🗄️ Load From DB (Existing Site)":
            if site_found:
                st.markdown("<br>", unsafe_allow_html=True)
                st.checkbox("", value=True, disabled=True, key="db_found_indicator", help="✅ Site found in database")
            else:
                st.markdown("<br>", unsafe_allow_html=True)
                st.caption("⚪ Not found")

    if "ipam_persisted_rows" not in st.session_state:
        st.session_state["ipam_persisted_rows"] = []

    # Sync editor deltas
    raw_rows = [dict(r) for r in st.session_state["ipam_persisted_rows"]]
    editor_state = st.session_state.get("ipam_data_editor_live", {})
    
    deleted_indices = set(editor_state.get("deleted_rows", []))
    if deleted_indices:
        raw_rows = [r for i, r in enumerate(raw_rows) if i not in deleted_indices]

    edited_cells = editor_state.get("edited_rows", {})
    for row_idx_str, changes in edited_cells.items():
        row_idx = int(row_idx_str)
        if row_idx < len(raw_rows):
            if "Role" in changes and "VLAN Name" not in changes:
                changes["VLAN Name"] = changes["Role"]
                if "VLAN Description" not in changes:
                    changes["VLAN Description"] = lookup_role_description(changes["Role"])
            # Apply Title Case formatting to Role and Description
            if "Role" in changes:
                changes["Role"] = to_title_case_preserve_acronyms(changes["Role"])
            if "VLAN Description" in changes:
                changes["VLAN Description"] = to_title_case_preserve_acronyms(changes["VLAN Description"])
            raw_rows[row_idx].update(changes)

    for new_r in editor_state.get("added_rows", []):
        r_name = new_r.get("Role", "")
        raw_rows.append({
            "VLAN ID": new_r.get("VLAN ID", None),
            "Role": to_title_case_preserve_acronyms(r_name),
            "VLAN Name": new_r.get("VLAN Name", r_name),
            "VLAN Description": to_title_case_preserve_acronyms(new_r.get("VLAN Description", lookup_role_description(r_name))),
            "Subnet (CIDR)": new_r.get("Subnet (CIDR)", "")
        })

    computed_rows = compute_chained_rows(supernet_in, raw_rows)
    st.session_state["ipam_persisted_rows"] = computed_rows

    allocated_subnets = []
    for r in computed_rows:
        sub_str = str(r.get("Subnet (CIDR)", "") or "").strip()
        if sub_str:
            allocated_subnets.append(sub_str)
        eval_res = evaluate_subnet_row(
            sub_str, 
            r.get("VLAN ID"), 
            r.get("Role", ""), 
            site_name, 
            supernet_in, 
            existing_prefixes
        )
        r["Usable Range"] = eval_res["usable_range"]
        r["Status"] = eval_res["status"]
        r["Prefix Description"] = eval_res["desc"]

    # Real-time Available Subnets and Capacity
    with cap_placeholder.container():
        if supernet_in and "/" in supernet_in:
            try:
                sup_net = ipaddress.ip_network(supernet_in, strict=False)
                sup_range = calculate_ip_range_str(sup_net)
                cap_matrix = calculate_remaining_subnets(supernet_in, allocated_subnets)
                cap_str = f"**Available:** `{cap_matrix['/24']}x /24` | `{cap_matrix['/25']}x /25` | `{cap_matrix['/26']}x /26` | `{cap_matrix['/27']}x /27`"
                st.markdown(f"📍 **Site Subnet:** `{sup_range}`")
                st.caption(cap_str)
            except ValueError:
                st.caption("⚠️ Invalid CIDR format")

    TABLE_COLS = [
        "VLAN ID", "Role", "VLAN Name", "VLAN Description", 
        "Suggest Subnet", "Subnet (CIDR)", "Usable Range", "Status", "Prefix Description"
    ]
    if computed_rows:
        df_init = pd.DataFrame(computed_rows)[TABLE_COLS]
    else:
        df_init = pd.DataFrame(columns=TABLE_COLS)

    edited_df = st.data_editor(
        df_init,
        width="stretch",
        num_rows="dynamic",
        key="ipam_data_editor_live",
        column_config={
            "VLAN ID": st.column_config.NumberColumn("VLAN ID", step=1, required=True),
            "Role": st.column_config.TextColumn("Role", help="VLAN Role. Auto-sets VLAN Name & Description via DB lookup."),
            "VLAN Name": st.column_config.TextColumn("VLAN Name", help="VLAN Name in NetBox. Defaults to Role, or editable."),
            "VLAN Description": st.column_config.TextColumn("VLAN Description", help="VLAN Description. Auto-looked up from DB or editable."),
            "Suggest Subnet": st.column_config.TextColumn("Suggest Subnet", help="Calculated next available network IP ID.", disabled=True),
            "Subnet (CIDR)": st.column_config.TextColumn("Subnet (CIDR)", help="Type subnet CIDR (e.g. 10.113.252.0/23) and hit Enter."),
            "Usable Range": st.column_config.TextColumn("Usable Range", help="Calculated usable host IP range.", disabled=True),
            "Status": st.column_config.TextColumn("Status", help="Collision & Database usage status.", disabled=True),
            "Prefix Description": st.column_config.TextColumn("Prefix Description", help="Calculated NetBox prefix description.", disabled=True)
        }
    )

    # 4. NetBox Bulk-Import CSV Copy Cards & Scope ID Notification
    st.markdown("---")
    st.markdown("### 📋 NetBox Bulk-Import CSV Generators")

    if not scope_id:
        st.warning("⚠️ **Notice:** Scope ID is empty. Export NetBox Sites CSV to discover Scope ID, or enter it manually.", icon="⚠️")

    display_site = display_site_name or "Site"
    csv_site = generate_netbox_site_csv(display_site)
    csv_group = generate_netbox_vlan_group_csv(display_site, scope_id)
    csv_vlans = generate_netbox_vlans_csv(display_site, computed_rows)
    csv_prefixes = generate_netbox_prefixes_csv(display_site, scope_id, supernet_in, computed_rows)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**1. Import Site (`dcim.site`)**")
        st.code(csv_site, language="csv")

        st.markdown("**3. Import VLANs (`ipam.vlan`)** *(Assigned IP subnets only)*")
        st.code(csv_vlans, language="csv")

    with c2:
        st.markdown("**2. Import VLAN Group (`ipam.vlangroup`)**")
        st.code(csv_group, language="csv")

        st.markdown("**4. Import Prefixes (`ipam.prefix`)**")
        st.code(csv_prefixes, language="csv")